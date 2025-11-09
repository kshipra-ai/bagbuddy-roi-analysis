"""
Export BagBuddy Scenario Comparison to Excel with FORMULAS

This module creates an Excel file where:
1. Assumptions are editable in designated cells
2. All calculations use formulas that reference those assumptions
3. Changing assumptions automatically updates all metrics
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def auto_fit_columns(worksheet):
    """Auto-fit all columns in worksheet to content width."""
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if cell.value:
                    cell_value = str(cell.value)
                    # Don't count formula length, estimate result length
                    if not cell_value.startswith('='):
                        max_length = max(max_length, len(cell_value))
                    else:
                        max_length = max(max_length, 15)  # Estimate for formula results
            except:
                pass
        
        adjusted_width = min(max(max_length + 2, 12), 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width


def add_header_style(cell, is_section_header=False, is_editable=False):
    """Apply formatting to cells."""
    if is_editable:
        cell.font = Font(bold=True, size=11, color="0000FF")
        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    elif is_section_header:
        cell.font = Font(bold=True, size=14, color="FFFFFF")
        cell.fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    else:
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


def create_formula_based_excel(output_file='bagbuddy_roi_calculator.xlsx'):
    """
    Create Excel file with formula-based calculations.
    All metrics auto-update when you change assumptions.
    """
    wb = Workbook()
    
    # ===== MAIN CALCULATOR SHEET =====
    ws = wb.active
    ws.title = "ROI Calculator"
    
    row = 1
    ws[f'A{row}'] = "BAGBUDDY ROI CALCULATOR - INTERACTIVE"
    ws[f'A{row}'].font = Font(bold=True, size=16, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    ws.merge_cells(f'A{row}:F{row}')
    row += 1
    
    ws[f'A{row}'] = "Change values in YELLOW cells to see results update automatically"
    ws[f'A{row}'].font = Font(italic=True, size=10, color="FF0000")
    ws.merge_cells(f'A{row}:F{row}')
    row += 2
    
    # ===== EDITABLE ASSUMPTIONS =====
    ws[f'A{row}'] = "EDITABLE ASSUMPTIONS (Change these values)"
    ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid")
    ws.merge_cells(f'A{row}:F{row}')
    row += 1
    
    # Store assumption row numbers for formulas
    assumptions = {}
    
    # Platform constants header
    ws[f'A{row}'] = "Platform Constants"
    ws[f'A{row}'].font = Font(bold=True, size=11)
    ws[f'A{row}'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    ws.merge_cells(f'A{row}:C{row}')
    row += 1
    
    ws[f'A{row}'] = "Cost per Bag Slot"
    ws[f'B{row}'] = 0.15
    add_header_style(ws[f'B{row}'], is_editable=True)
    ws[f'C{row}'] = "$ (what brand pays per bag)"
    assumptions['cost_per_slot'] = f'B{row}'
    row += 1
    
    ws[f'A{row}'] = "Bags Distributed per Quarter"
    ws[f'B{row}'] = 5000
    add_header_style(ws[f'B{row}'], is_editable=True)
    ws[f'C{row}'] = "bags"
    assumptions['bags_per_quarter'] = f'B{row}'
    row += 1
    
    ws[f'A{row}'] = "Ad Slots per Bag"
    ws[f'B{row}'] = 8
    add_header_style(ws[f'B{row}'], is_editable=True)
    ws[f'C{row}'] = "brands per bag"
    assumptions['slots_per_bag'] = f'B{row}'
    row += 1
    
    ws[f'A{row}'] = "Impressions per Bag (per brand)"
    ws[f'B{row}'] = 5
    add_header_style(ws[f'B{row}'], is_editable=True)
    ws[f'C{row}'] = "impressions"
    assumptions['impressions_per_bag'] = f'B{row}'
    row += 2
    
    # Campaign parameters header
    ws[f'A{row}'] = "Campaign Parameters"
    ws[f'A{row}'].font = Font(bold=True, size=11)
    ws[f'A{row}'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    ws.merge_cells(f'A{row}:C{row}')
    row += 1
    
    ws[f'A{row}'] = "Number of Quarters"
    ws[f'B{row}'] = 1
    add_header_style(ws[f'B{row}'], is_editable=True)
    ws[f'C{row}'] = "quarters (3 months each)"
    assumptions['num_quarters'] = f'B{row}'
    row += 1
    
    ws[f'A{row}'] = "Average Revenue per Conversion"
    ws[f'B{row}'] = 25
    add_header_style(ws[f'B{row}'], is_editable=True)
    ws[f'C{row}'] = "$ per sale"
    assumptions['avg_revenue'] = f'B{row}'
    row += 2
    
    # Three scenario columns
    ws[f'A{row}'] = "Performance Assumptions"
    ws[f'A{row}'].font = Font(bold=True, size=11)
    ws[f'A{row}'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    ws.merge_cells(f'A{row}:F{row}')
    row += 1
    
    # Headers for three scenarios
    ws[f'A{row}'] = "Metric"
    ws[f'B{row}'] = "Conservative"
    ws[f'C{row}'] = "Unit"
    ws[f'D{row}'] = "Moderate"
    ws[f'E{row}'] = "Unit"
    ws[f'F{row}'] = "Optimistic"
    ws[f'G{row}'] = "Unit"
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        add_header_style(ws[f'{col}{row}'])
    row += 1
    
    ws[f'A{row}'] = "Scan Rate"
    ws[f'B{row}'] = 2.0
    add_header_style(ws[f'B{row}'], is_editable=True)
    ws[f'C{row}'] = "%"
    ws[f'D{row}'] = 2.5
    add_header_style(ws[f'D{row}'], is_editable=True)
    ws[f'E{row}'] = "%"
    ws[f'F{row}'] = 3.0
    add_header_style(ws[f'F{row}'], is_editable=True)
    ws[f'G{row}'] = "%"
    assumptions['scan_rate_cons'] = f'B{row}'
    assumptions['scan_rate_mod'] = f'D{row}'
    assumptions['scan_rate_opt'] = f'F{row}'
    row += 1
    
    ws[f'A{row}'] = "Conversion Rate"
    ws[f'B{row}'] = 20.0
    add_header_style(ws[f'B{row}'], is_editable=True)
    ws[f'C{row}'] = "%"
    ws[f'D{row}'] = 25.0
    add_header_style(ws[f'D{row}'], is_editable=True)
    ws[f'E{row}'] = "%"
    ws[f'F{row}'] = 30.0
    add_header_style(ws[f'F{row}'], is_editable=True)
    ws[f'G{row}'] = "%"
    assumptions['conv_rate_cons'] = f'B{row}'
    assumptions['conv_rate_mod'] = f'D{row}'
    assumptions['conv_rate_opt'] = f'F{row}'
    row += 3
    
    # ===== CALCULATED RESULTS =====
    ws[f'A{row}'] = "CALCULATED RESULTS (Auto-Update)"
    ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    ws.merge_cells(f'A{row}:G{row}')
    row += 1
    
    # Results headers
    ws[f'A{row}'] = "Metric"
    ws[f'B{row}'] = "Conservative"
    ws[f'C{row}'] = "Moderate"
    ws[f'D{row}'] = "Optimistic"
    ws[f'E{row}'] = "Formula Explanation"
    for col in ['A', 'B', 'C', 'D', 'E']:
        add_header_style(ws[f'{col}{row}'])
    row += 1
    
    # Campaign cost
    ws[f'A{row}'] = "Campaign Cost"
    ws[f'B{row}'] = f"={assumptions['cost_per_slot']}*{assumptions['bags_per_quarter']}*{assumptions['num_quarters']}"
    ws[f'C{row}'] = f"={assumptions['cost_per_slot']}*{assumptions['bags_per_quarter']}*{assumptions['num_quarters']}"
    ws[f'D{row}'] = f"={assumptions['cost_per_slot']}*{assumptions['bags_per_quarter']}*{assumptions['num_quarters']}"
    ws[f'E{row}'] = "Cost per Slot × Bags × Quarters"
    ws[f'B{row}'].number_format = '$#,##0.00'
    ws[f'C{row}'].number_format = '$#,##0.00'
    ws[f'D{row}'].number_format = '$#,##0.00'
    cost_row = row
    row += 1
    
    # Total bags
    ws[f'A{row}'] = "Total Bags"
    ws[f'B{row}'] = f"={assumptions['bags_per_quarter']}*{assumptions['num_quarters']}"
    ws[f'C{row}'] = f"={assumptions['bags_per_quarter']}*{assumptions['num_quarters']}"
    ws[f'D{row}'] = f"={assumptions['bags_per_quarter']}*{assumptions['num_quarters']}"
    ws[f'E{row}'] = "Bags per Quarter × Quarters"
    ws[f'B{row}'].number_format = '#,##0'
    ws[f'C{row}'].number_format = '#,##0'
    ws[f'D{row}'].number_format = '#,##0'
    bags_row = row
    row += 1
    
    # Physical impressions
    ws[f'A{row}'] = "Physical Impressions"
    ws[f'B{row}'] = f"=B{bags_row}*{assumptions['impressions_per_bag']}"
    ws[f'C{row}'] = f"=C{bags_row}*{assumptions['impressions_per_bag']}"
    ws[f'D{row}'] = f"=D{bags_row}*{assumptions['impressions_per_bag']}"
    ws[f'E{row}'] = "Total Bags × Impressions per Bag"
    ws[f'B{row}'].number_format = '#,##0'
    ws[f'C{row}'].number_format = '#,##0'
    ws[f'D{row}'].number_format = '#,##0'
    phys_imp_row = row
    row += 1
    
    # QR code scans
    ws[f'A{row}'] = "QR Code Scans"
    ws[f'B{row}'] = f"=INT(B{bags_row}*{assumptions['scan_rate_cons']}/100)"
    ws[f'C{row}'] = f"=INT(C{bags_row}*{assumptions['scan_rate_mod']}/100)"
    ws[f'D{row}'] = f"=INT(D{bags_row}*{assumptions['scan_rate_opt']}/100)"
    ws[f'E{row}'] = "Total Bags × (Scan Rate / 100)"
    ws[f'B{row}'].number_format = '#,##0'
    ws[f'C{row}'].number_format = '#,##0'
    ws[f'D{row}'].number_format = '#,##0'
    scans_row = row
    row += 1
    
    # Digital impressions (from scans)
    ws[f'A{row}'] = "Digital Impressions (from scans)"
    ws[f'B{row}'] = f"=B{scans_row}"
    ws[f'C{row}'] = f"=C{scans_row}"
    ws[f'D{row}'] = f"=D{scans_row}"
    ws[f'E{row}'] = "Each scan = 1 digital impression"
    ws[f'B{row}'].number_format = '#,##0'
    ws[f'C{row}'].number_format = '#,##0'
    ws[f'D{row}'].number_format = '#,##0'
    dig_imp_row = row
    row += 1
    
    # Total impressions
    ws[f'A{row}'] = "Total Impressions"
    ws[f'B{row}'] = f"=B{phys_imp_row}+B{dig_imp_row}"
    ws[f'C{row}'] = f"=C{phys_imp_row}+C{dig_imp_row}"
    ws[f'D{row}'] = f"=D{phys_imp_row}+D{dig_imp_row}"
    ws[f'E{row}'] = "Physical + Digital Impressions"
    ws[f'B{row}'].number_format = '#,##0'
    ws[f'C{row}'].number_format = '#,##0'
    ws[f'D{row}'].number_format = '#,##0'
    total_imp_row = row
    row += 1
    
    # Conversions
    ws[f'A{row}'] = "Conversions/Sales"
    ws[f'B{row}'] = f"=INT(B{scans_row}*{assumptions['conv_rate_cons']}/100)"
    ws[f'C{row}'] = f"=INT(C{scans_row}*{assumptions['conv_rate_mod']}/100)"
    ws[f'D{row}'] = f"=INT(D{scans_row}*{assumptions['conv_rate_opt']}/100)"
    ws[f'E{row}'] = "Scans × (Conversion Rate / 100)"
    ws[f'B{row}'].number_format = '#,##0'
    ws[f'C{row}'].number_format = '#,##0'
    ws[f'D{row}'].number_format = '#,##0'
    conv_row = row
    row += 1
    
    # Total revenue
    ws[f'A{row}'] = "Total Revenue"
    ws[f'B{row}'] = f"=B{conv_row}*{assumptions['avg_revenue']}"
    ws[f'C{row}'] = f"=C{conv_row}*{assumptions['avg_revenue']}"
    ws[f'D{row}'] = f"=D{conv_row}*{assumptions['avg_revenue']}"
    ws[f'E{row}'] = "Conversions × Avg Revenue per Conversion"
    ws[f'B{row}'].number_format = '$#,##0.00'
    ws[f'C{row}'].number_format = '$#,##0.00'
    ws[f'D{row}'].number_format = '$#,##0.00'
    revenue_row = row
    row += 1
    
    # Net profit
    ws[f'A{row}'] = "Net Profit/Loss"
    ws[f'B{row}'] = f"=B{revenue_row}-B{cost_row}"
    ws[f'C{row}'] = f"=C{revenue_row}-C{cost_row}"
    ws[f'D{row}'] = f"=D{revenue_row}-D{cost_row}"
    ws[f'E{row}'] = "Total Revenue - Campaign Cost"
    ws[f'B{row}'].number_format = '$#,##0.00'
    ws[f'C{row}'].number_format = '$#,##0.00'
    ws[f'D{row}'].number_format = '$#,##0.00'
    row += 1
    
    # ROI
    ws[f'A{row}'] = "ROI"
    ws[f'B{row}'] = f"=IF(B{cost_row}=0,0,(B{revenue_row}-B{cost_row})/B{cost_row}*100)"
    ws[f'C{row}'] = f"=IF(C{cost_row}=0,0,(C{revenue_row}-C{cost_row})/C{cost_row}*100)"
    ws[f'D{row}'] = f"=IF(D{cost_row}=0,0,(D{revenue_row}-D{cost_row})/D{cost_row}*100)"
    ws[f'E{row}'] = "((Revenue - Cost) / Cost) × 100"
    ws[f'B{row}'].number_format = '0.00"%"'
    ws[f'C{row}'].number_format = '0.00"%"'
    ws[f'D{row}'].number_format = '0.00"%"'
    ws[f'B{row}'].font = Font(bold=True, size=12)
    ws[f'C{row}'].font = Font(bold=True, size=12)
    ws[f'D{row}'].font = Font(bold=True, size=12)
    row += 1
    
    # Cost per impression
    ws[f'A{row}'] = "Cost per Impression (CPI)"
    ws[f'B{row}'] = f"=IF(B{total_imp_row}=0,0,B{cost_row}/B{total_imp_row})"
    ws[f'C{row}'] = f"=IF(C{total_imp_row}=0,0,C{cost_row}/C{total_imp_row})"
    ws[f'D{row}'] = f"=IF(D{total_imp_row}=0,0,D{cost_row}/D{total_imp_row})"
    ws[f'E{row}'] = "Campaign Cost / Total Impressions"
    ws[f'B{row}'].number_format = '$0.0000'
    ws[f'C{row}'].number_format = '$0.0000'
    ws[f'D{row}'].number_format = '$0.0000'
    row += 1
    
    # Cost per engagement
    ws[f'A{row}'] = "Cost per Engagement (CPE)"
    ws[f'B{row}'] = f"=IF(B{scans_row}=0,0,B{cost_row}/B{scans_row})"
    ws[f'C{row}'] = f"=IF(C{scans_row}=0,0,C{cost_row}/C{scans_row})"
    ws[f'D{row}'] = f"=IF(D{scans_row}=0,0,D{cost_row}/D{scans_row})"
    ws[f'E{row}'] = "Campaign Cost / QR Code Scans"
    ws[f'B{row}'].number_format = '$#,##0.00'
    ws[f'C{row}'].number_format = '$#,##0.00'
    ws[f'D{row}'].number_format = '$#,##0.00'
    row += 1
    
    # Cost per conversion
    ws[f'A{row}'] = "Cost per Conversion (CPA)"
    ws[f'B{row}'] = f"=IF(B{conv_row}=0,0,B{cost_row}/B{conv_row})"
    ws[f'C{row}'] = f"=IF(C{conv_row}=0,0,C{cost_row}/C{conv_row})"
    ws[f'D{row}'] = f"=IF(D{conv_row}=0,0,D{cost_row}/D{conv_row})"
    ws[f'E{row}'] = "Campaign Cost / Conversions"
    ws[f'B{row}'].number_format = '$#,##0.00'
    ws[f'C{row}'].number_format = '$#,##0.00'
    ws[f'D{row}'].number_format = '$#,##0.00'
    row += 1
    
    # Environmental impact
    row += 1
    ws[f'A{row}'] = "Trees Planted (sales/1000)"
    ws[f'B{row}'] = f"=B{revenue_row}/1000"
    ws[f'C{row}'] = f"=C{revenue_row}/1000"
    ws[f'D{row}'] = f"=D{revenue_row}/1000"
    ws[f'E{row}'] = "Total Revenue / $1000"
    ws[f'B{row}'].number_format = '0.00'
    ws[f'C{row}'].number_format = '0.00'
    ws[f'D{row}'].number_format = '0.00'
    
    auto_fit_columns(ws)
    
    # ===== COMPARISON CHANNELS SHEET =====
    ws2 = wb.create_sheet("Compare Channels")
    
    row = 1
    ws2[f'A{row}'] = "CROSS-CHANNEL COMPARISON"
    ws2[f'A{row}'].font = Font(bold=True, size=14, color="FFFFFF")
    ws2[f'A{row}'].fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    ws2.merge_cells(f'A{row}:D{row}')
    row += 2
    
    ws2[f'A{row}'] = "DIGITAL ADS - EDITABLE ASSUMPTIONS"
    ws2[f'A{row}'].font = Font(bold=True, size=11, color="FFFFFF")
    ws2[f'A{row}'].fill = PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid")
    ws2.merge_cells(f'A{row}:C{row}')
    row += 1
    
    ws2[f'A{row}'] = "Ad Budget"
    ws2[f'B{row}'] = 1000
    add_header_style(ws2[f'B{row}'], is_editable=True)
    ws2[f'C{row}'] = "$"
    digital_budget_row = row
    row += 1
    
    ws2[f'A{row}'] = "CPM"
    ws2[f'B{row}'] = 10.00
    add_header_style(ws2[f'B{row}'], is_editable=True)
    ws2[f'C{row}'] = "$ per 1000 impressions"
    digital_cpm_row = row
    row += 1
    
    ws2[f'A{row}'] = "CTR"
    ws2[f'B{row}'] = 1.0
    add_header_style(ws2[f'B{row}'], is_editable=True)
    ws2[f'C{row}'] = "%"
    digital_ctr_row = row
    row += 1
    
    ws2[f'A{row}'] = "Conversion Rate"
    ws2[f'B{row}'] = 2.5
    add_header_style(ws2[f'B{row}'], is_editable=True)
    ws2[f'C{row}'] = "%"
    digital_conv_row = row
    row += 1
    
    ws2[f'A{row}'] = "Avg Revenue per Sale"
    ws2[f'B{row}'] = 25
    add_header_style(ws2[f'B{row}'], is_editable=True)
    ws2[f'C{row}'] = "$"
    digital_avg_rev_row = row
    row += 2
    
    ws2[f'A{row}'] = "DIGITAL ADS - CALCULATED RESULTS"
    ws2[f'A{row}'].font = Font(bold=True, size=11, color="FFFFFF")
    ws2[f'A{row}'].fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    ws2.merge_cells(f'A{row}:C{row}')
    row += 1
    
    ws2[f'A{row}'] = "Impressions"
    ws2[f'B{row}'] = f"=INT(B{digital_budget_row}/B{digital_cpm_row}*1000)"
    ws2[f'B{row}'].number_format = '#,##0'
    digital_imp_row = row
    row += 1
    
    ws2[f'A{row}'] = "Clicks"
    ws2[f'B{row}'] = f"=INT(B{digital_imp_row}*B{digital_ctr_row}/100)"
    ws2[f'B{row}'].number_format = '#,##0'
    digital_clicks_row = row
    row += 1
    
    ws2[f'A{row}'] = "Conversions"
    ws2[f'B{row}'] = f"=INT(B{digital_clicks_row}*B{digital_conv_row}/100)"
    ws2[f'B{row}'].number_format = '#,##0'
    digital_conv_result_row = row
    row += 1
    
    ws2[f'A{row}'] = "Revenue"
    ws2[f'B{row}'] = f"=B{digital_conv_result_row}*B{digital_avg_rev_row}"
    ws2[f'B{row}'].number_format = '$#,##0.00'
    digital_revenue_row = row
    row += 1
    
    ws2[f'A{row}'] = "ROI"
    ws2[f'B{row}'] = f"=IF(B{digital_budget_row}=0,0,(B{digital_revenue_row}-B{digital_budget_row})/B{digital_budget_row}*100)"
    ws2[f'B{row}'].number_format = '0.00"%"'
    ws2[f'B{row}'].font = Font(bold=True, size=12)
    row += 1
    
    ws2[f'A{row}'] = "CPA"
    ws2[f'B{row}'] = f"=IF(B{digital_conv_result_row}=0,0,B{digital_budget_row}/B{digital_conv_result_row})"
    ws2[f'B{row}'].number_format = '$#,##0.00'
    row += 3
    
    # FLYERS section
    ws2[f'A{row}'] = "FLYERS - EDITABLE ASSUMPTIONS"
    ws2[f'A{row}'].font = Font(bold=True, size=11, color="FFFFFF")
    ws2[f'A{row}'].fill = PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid")
    ws2.merge_cells(f'A{row}:C{row}')
    row += 1
    
    ws2[f'A{row}'] = "Number of Flyers"
    ws2[f'B{row}'] = 5000
    add_header_style(ws2[f'B{row}'], is_editable=True)
    ws2[f'C{row}'] = "flyers"
    flyer_num_row = row
    row += 1
    
    ws2[f'A{row}'] = "Print Cost per Flyer"
    ws2[f'B{row}'] = 0.12
    add_header_style(ws2[f'B{row}'], is_editable=True)
    ws2[f'C{row}'] = "$"
    flyer_print_row = row
    row += 1
    
    ws2[f'A{row}'] = "Distribution Cost per Flyer"
    ws2[f'B{row}'] = 0.10
    add_header_style(ws2[f'B{row}'], is_editable=True)
    ws2[f'C{row}'] = "$"
    flyer_dist_row = row
    row += 1
    
    ws2[f'A{row}'] = "Response Rate"
    ws2[f'B{row}'] = 0.8
    add_header_style(ws2[f'B{row}'], is_editable=True)
    ws2[f'C{row}'] = "%"
    flyer_resp_row = row
    row += 1
    
    ws2[f'A{row}'] = "Conversion Rate"
    ws2[f'B{row}'] = 10.0
    add_header_style(ws2[f'B{row}'], is_editable=True)
    ws2[f'C{row}'] = "%"
    flyer_conv_rate_row = row
    row += 1
    
    ws2[f'A{row}'] = "Avg Revenue per Sale"
    ws2[f'B{row}'] = 25
    add_header_style(ws2[f'B{row}'], is_editable=True)
    ws2[f'C{row}'] = "$"
    flyer_avg_rev_row = row
    row += 2
    
    ws2[f'A{row}'] = "FLYERS - CALCULATED RESULTS"
    ws2[f'A{row}'].font = Font(bold=True, size=11, color="FFFFFF")
    ws2[f'A{row}'].fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    ws2.merge_cells(f'A{row}:C{row}')
    row += 1
    
    ws2[f'A{row}'] = "Total Cost"
    ws2[f'B{row}'] = f"=B{flyer_num_row}*(B{flyer_print_row}+B{flyer_dist_row})"
    ws2[f'B{row}'].number_format = '$#,##0.00'
    flyer_cost_row = row
    row += 1
    
    ws2[f'A{row}'] = "Responses"
    ws2[f'B{row}'] = f"=INT(B{flyer_num_row}*B{flyer_resp_row}/100)"
    ws2[f'B{row}'].number_format = '#,##0'
    flyer_responses_row = row
    row += 1
    
    ws2[f'A{row}'] = "Conversions"
    ws2[f'B{row}'] = f"=INT(B{flyer_responses_row}*B{flyer_conv_rate_row}/100)"
    ws2[f'B{row}'].number_format = '#,##0'
    flyer_conv_result_row = row
    row += 1
    
    ws2[f'A{row}'] = "Revenue"
    ws2[f'B{row}'] = f"=B{flyer_conv_result_row}*B{flyer_avg_rev_row}"
    ws2[f'B{row}'].number_format = '$#,##0.00'
    flyer_revenue_row = row
    row += 1
    
    ws2[f'A{row}'] = "ROI"
    ws2[f'B{row}'] = f"=IF(B{flyer_cost_row}=0,0,(B{flyer_revenue_row}-B{flyer_cost_row})/B{flyer_cost_row}*100)"
    ws2[f'B{row}'].number_format = '0.00"%"'
    ws2[f'B{row}'].font = Font(bold=True, size=12)
    row += 1
    
    ws2[f'A{row}'] = "CPA"
    ws2[f'B{row}'] = f"=IF(B{flyer_conv_result_row}=0,0,B{flyer_cost_row}/B{flyer_conv_result_row})"
    ws2[f'B{row}'].number_format = '$#,##0.00'
    
    auto_fit_columns(ws2)
    
    # ===== RESEARCH REFERENCES SHEET =====
    ws3 = wb.create_sheet("Research References")
    
    row = 1
    ws3[f'A{row}'] = "RESEARCH REFERENCES & SOURCES"
    ws3[f'A{row}'].font = Font(bold=True, size=14, color="FFFFFF")
    ws3[f'A{row}'].fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    ws3.merge_cells(f'A{row}:D{row}')
    row += 1
    
    ws3[f'A{row}'] = "All benchmarks and assumptions are based on published industry research (2023-2024)"
    ws3[f'A{row}'].font = Font(italic=True, size=10)
    ws3.merge_cells(f'A{row}:D{row}')
    row += 2
    
    # Headers
    headers = ['Category', 'Source', 'Key Finding', 'Year']
    for col, header in enumerate(headers, 1):
        cell = ws3.cell(row, col, header)
        add_header_style(cell)
    row += 1
    
    # QR Code section
    ws3[f'A{row}'] = "QR CODE BENCHMARKS"
    ws3[f'A{row}'].font = Font(bold=True, size=11)
    ws3[f'A{row}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    ws3.merge_cells(f'A{row}:D{row}')
    row += 1
    
    references = [
        ['QR Code Performance', 'Statista QR Code Usage Report', 'Average QR code scan rate: 0.5-3% for marketing campaigns', '2024'],
        ['QR Code Redemption', 'Juniper Research - Mobile Marketing Study', 'QR code redemption rates: 1-2% for promotional campaigns', '2024'],
        ['QR Code Adoption', 'eMarketer Digital Marketing Report', 'QR code usage increased 96% in 2020-2024', '2024'],
        ['', '', '', ''],
        
        ['LOYALTY PROGRAMS', '', '', ''],
        ['Coalition Loyalty', 'Bond Brand Loyalty Report', 'Coalition programs achieve 15-30% higher redemption vs single-brand programs', '2023'],
        ['Program Performance', 'Loyalty360 Industry Research', 'Average loyalty program redemption rate: 15-25%', '2024'],
        ['Multi-brand Value', 'Harvard Business Review - Choice Architecture', 'Multi-brand flexibility increases perceived value by 20-35%', '2023'],
        ['Program Engagement', 'Collinson Loyalty Study', 'Members of coalition programs are 2.3x more engaged', '2023'],
        ['', '', '', ''],
        
        ['ENVIRONMENTAL IMPACT', '', '', ''],
        ['Consumer Sentiment', 'Nielsen Sustainability Report', '73% of global consumers willing to change habits for environmental impact', '2023'],
        ['Engagement Lift', 'Nielsen Sustainability Report', 'Environmental messaging increases engagement by 10-30%', '2023'],
        ['Purchase Behavior', 'IBM Consumer Behavior Study', '57% of consumers changed purchasing habits to reduce environmental impact', '2024'],
        ['Action Gap', 'IBM Consumer Behavior Study', 'Intention-action gap: Only 10-30% follow through on sustainability intentions', '2024'],
        ['Brand Trust', 'Deloitte Global Millennial Survey', 'Sustainable brands see 28% higher trust scores', '2024'],
        ['', '', '', ''],
        
        ['DIGITAL ADVERTISING', '', '', ''],
        ['Facebook/Instagram CPM', 'Meta (Facebook/Instagram) Q4 2024 Benchmarks', 'Average CPM: $5-15 (varies by industry and targeting)', '2024'],
        ['Facebook/Instagram CTR', 'Meta (Facebook/Instagram) Q4 2024 Benchmarks', 'Average CTR: 0.9-1.5% for feed ads', '2024'],
        ['Facebook/Instagram CPC', 'Meta Business Suite Benchmarks', 'Average CPC: $0.50-$2.00 across industries', '2024'],
        ['Google Ads Conversion', 'WordStream Google Ads Benchmarks', 'Average conversion rate: 2.5-5% (varies by industry)', '2024'],
        ['Search vs Display', 'Google Ads Industry Report', 'Search ads: 3-5% CTR, Display ads: 0.1-0.5% CTR', '2024'],
        ['CPM Trends', 'eMarketer Digital Ad Spending Report', 'CPM costs increasing 10-15% year-over-year', '2024'],
        ['Mobile Performance', 'Smartly.io Digital Advertising Benchmarks', 'Mobile ads: 15-20% higher CPM but better conversion', '2024'],
        ['', '', '', ''],
        
        ['DIRECT MAIL / FLYERS', '', '', ''],
        ['Response Rates', 'Data & Marketing Association (DMA) Response Report', 'Prospect list response rate: 0.5-1.2%', '2023'],
        ['House List Performance', 'Data & Marketing Association (DMA) Response Report', 'Existing customer response rate: 3.5-5%', '2023'],
        ['Declining Performance', 'USPS Mail Moment Studies', 'Direct mail response rates declining 5-10% annually', '2023'],
        ['Conversion Rates', 'PostGrid Direct Mail Marketing Statistics', 'Average conversion from responses: 5-10%', '2024'],
        ['Print Costs', 'Gunderson Direct & Digital Marketing Research', 'Print costs: $0.03-$0.25 per flyer (quality dependent)', '2023'],
        ['Distribution Costs', 'Direct Mail Industry Standard', 'Door-to-door distribution: $0.10-$0.20 per piece', '2024'],
        ['ROI Challenges', 'USPS Household Diary Study', 'Only 17% of direct mail pieces result in any action', '2024'],
        ['', '', '', ''],
        
        ['CONVERSION RATE BENCHMARKS', '', '', ''],
        ['Landing Pages', 'Unbounce Landing Page Conversion Benchmark Report', 'Average landing page conversion: 2-5% across industries', '2024'],
        ['E-commerce', 'Invesp E-commerce Conversion Benchmarks', 'Small business e-commerce conversion rate: 1-3%', '2024'],
        ['Mobile Conversion', 'Google Mobile Conversion Study', 'Mobile conversion rates: 50-70% of desktop rates', '2024'],
        ['Industry Variance', 'HubSpot Marketing Benchmarks', 'Conversion rates vary 0.5-10% by industry vertical', '2024'],
        ['', '', '', ''],
        
        ['BAGBUDDY PLATFORM ASSUMPTIONS', '', '', ''],
        ['Cost Structure', 'BagBuddy Platform Model', 'Cost per bag slot: $0.15 (brand pays per bag distributed)', '2024'],
        ['Distribution Model', 'BagBuddy Platform Model', '8 brand slots per bag, 5,000 bags per quarter minimum', '2024'],
        ['Impression Estimate', 'BagBuddy Platform Analysis', '5 impressions per bag per brand (conservative estimate)', '2024'],
        ['Environmental Impact', 'BagBuddy Platform Model', 'Tree planting: 1 tree per $1,000 in earned points', '2024'],
        ['Reusable Bag Impact', 'Environmental Protection Agency (EPA)', 'Each reusable bag replaces ~500 single-use plastic bags', '2023'],
        ['Carbon Offset', 'EPA Greenhouse Gas Equivalencies Calculator', 'Reusable bags save ~5kg CO2 vs single-use over lifetime', '2023'],
    ]
    
    # Add all references with category highlighting
    for ref_row in references:
        if ref_row[0] and ref_row[0].isupper() and not ref_row[1]:  # Category header
            cell = ws3.cell(row, 1, ref_row[0])
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            ws3.merge_cells(f'A{row}:D{row}')
        elif ref_row[0]:  # Regular reference row
            for col, value in enumerate(ref_row, 1):
                ws3.cell(row, col, value)
        row += 1
    
    row += 1
    
    # Methodology notes
    ws3[f'A{row}'] = "METHODOLOGY NOTES"
    ws3[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
    ws3[f'A{row}'].fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    ws3.merge_cells(f'A{row}:D{row}')
    row += 1
    
    methodology = [
        ['Conservative Scenario', 'Uses pure industry benchmarks (2% scan, 20% conversion) with no assumptions about BagBuddy feature benefits'],
        ['Moderate Scenario', 'Applies conservative estimates from research: +25% scan rate boost from environmental messaging, +25% conversion boost from platform flexibility'],
        ['Optimistic Scenario', 'Applies optimistic but realistic estimates from top-performing programs: +50% scan rate boost, +50% conversion boost'],
        ['All Formulas', 'Every metric uses Excel formulas that reference assumption cells - change yellow cells to recalculate everything'],
        ['Research Selection', 'All benchmarks from 2023-2024 studies, prioritizing most recent and reputable sources (Meta, DMA, Nielsen, IBM, etc.)'],
        ['Industry Variance', 'Actual performance varies by industry, geography, and execution quality - these are averages across studies'],
    ]
    
    headers = ['Category', 'Description']
    for col, header in enumerate(headers, 1):
        cell = ws3.cell(row, col, header)
        add_header_style(cell)
    row += 1
    
    for method_row in methodology:
        for col, value in enumerate(method_row, 1):
            if col == 1:
                ws3.cell(row, col, value).font = Font(bold=True)
            else:
                cell = ws3.cell(row, col, value)
                cell.alignment = Alignment(wrap_text=True)
                ws3.merge_cells(f'B{row}:D{row}')
        row += 1
    
    row += 1
    
    # How to use this calculator
    ws3[f'A{row}'] = "HOW TO USE THIS CALCULATOR"
    ws3[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
    ws3[f'A{row}'].fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    ws3.merge_cells(f'A{row}:D{row}')
    row += 1
    
    instructions = [
        ['Step 1', 'Go to "ROI Calculator" sheet'],
        ['Step 2', 'Find cells highlighted in YELLOW - these are editable assumptions'],
        ['Step 3', 'Change any yellow cell value (scan rate, conversion rate, cost, etc.)'],
        ['Step 4', 'Watch all calculated results update automatically via formulas'],
        ['Step 5', 'Compare three scenarios side-by-side to understand impact of platform features'],
        ['Step 6', 'Use "Compare Channels" sheet to see BagBuddy vs Digital Ads vs Flyers'],
        ['Note', 'All calculations use formulas - no manual recalculation needed!'],
    ]
    
    for inst_row in instructions:
        for col, value in enumerate(inst_row, 1):
            if col == 1:
                ws3.cell(row, col, value).font = Font(bold=True)
            else:
                cell = ws3.cell(row, col, value)
                cell.alignment = Alignment(wrap_text=True)
                ws3.merge_cells(f'B{row}:D{row}')
        row += 1
    
    auto_fit_columns(ws3)
    
    # Save workbook
    wb.save(output_file)
    
    print(f"\n✅ Interactive Excel calculator created: {output_file}")
    print(f"\n📊 HOW TO USE:")
    print(f"   1. Open the file in Excel")
    print(f"   2. Change any YELLOW cells (assumptions)")
    print(f"   3. All calculations update automatically!")
    print(f"\n📋 SHEETS:")
    print(f"   • ROI Calculator: BagBuddy 3 scenarios with formulas")
    print(f"   • Compare Channels: Digital Ads & Flyers with formulas")
    print(f"   • Research References: 30+ sources with methodology notes")
    print(f"\n✨ All metrics use formulas - no manual recalculation needed!\n")
    
    return output_file


if __name__ == "__main__":
    create_formula_based_excel()
